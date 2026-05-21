# ============================================================
# KANZLEI AI — DATEV & EXPORT SERVICE v1.0
# Datei: core/export_service.py
#
# Unterstützte Formate:
#   ✓ DATEV Buchungsstapel (CSV nach DATEV-Standard)
#   ✓ DATEV Stammdaten (Debitoren/Kreditoren)
#   ✓ ELSTER-kompatibles XML (Grundstruktur)
#   ✓ Excel-Report (openpyxl, formatiert)
#   ✓ CSV-Universal-Export
#   ✓ ZIP-Paket (alle Formate gebündelt)
# ============================================================

import csv
import json
import re
import xml.etree.ElementTree as ET
import zipfile
import io
import os
import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Any, Tuple

from core.datev_export_utils import (
    DATEV_BUCHUNG_SPALTEN,
    DATEV_CATEGORY_BUCHUNGSSTAPEL,
    DATEV_ENCODING,
    DATEV_VERSION,
    build_datev_buchungen,
    buchung_to_row,
    normalize_berater_nr,
    normalize_mandanten_nr,
    sanitize_datev_text,
    validate_datev_buchungsstapel_csv,
)

log = logging.getLogger("kanzlei_export")


# ============================================================
# DATEV BUCHUNGSSTAPEL (CSV)
# Offizielles DATEV-Format für den Import in DATEV Kanzlei-Rechnungswesen
# ============================================================

def export_datev_buchungsstapel(
    mandant: str,
    mandant_daten: Dict,
    aufgaben: List[Dict],
    berater_nr: str = "1234",
    mandanten_nr: str = "00000",
    wirtschaftsjahr: int = None,
) -> bytes:
    """
    Erstellt einen DATEV-kompatiblen Buchungsstapel (CSV).
    Format entspricht DATEV Datenformat Buchungsstapel v700.

    Hinweis: Echte Buchungsdaten (SKR03/SKR04 Konten) müssen
    aus einem Buchhaltungssystem kommen. Diese Funktion erstellt
    die korrekte Dateistruktur mit Beispieldaten.
    """
    jahr = wirtschaftsjahr or datetime.now().year
    jetzt = datetime.now()
    berater_nr = normalize_berater_nr(berater_nr)
    mandanten_nr = normalize_mandanten_nr(mandant, mandant_daten, mandanten_nr)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")

    bezeichnung = sanitize_datev_text(f"Kanzlei AI Export {mandant}", 60)

    writer.writerow([
        "EXTF",
        DATEV_VERSION,
        DATEV_CATEGORY_BUCHUNGSSTAPEL,
        "Buchungsstapel",
        "9",
        jetzt.strftime("%Y%m%d%H%M%S") + "000",
        "",
        "",
        "Kanzlei AI",
        "",
        berater_nr,
        mandanten_nr,
        f"{jahr}0101",
        "4",
        jetzt.strftime("%Y%m%d"),
        jetzt.strftime("%Y%m%d"),
        bezeichnung,
        "",
        "1",
        "0",
        "0",
        "EUR",
        "", "", "", "", "", "",
    ])

    writer.writerow(list(DATEV_BUCHUNG_SPALTEN))

    buchungen = build_datev_buchungen(mandant, mandant_daten, aufgaben, jetzt)
    for b in buchungen:
        writer.writerow(buchung_to_row(b))

    csv_bytes = output.getvalue().encode(DATEV_ENCODING, errors="replace")
    meta = validate_datev_buchungsstapel_csv(csv_bytes)
    if meta.get("warnings"):
        log.info("DATEV Export %s: %s", mandant, "; ".join(meta["warnings"]))
    return csv_bytes


# ============================================================
# DATEV STAMMDATEN (Debitoren)
# ============================================================

def export_datev_stammdaten(
    mandanten: Dict[str, Dict],
    berater_nr: str = "1234",
) -> bytes:
    """
    Erstellt DATEV-Stammdaten für alle Mandanten (Debitoren).
    Für den Import der Mandanten als Debitorenkonten in DATEV.
    """
    berater_nr = normalize_berater_nr(berater_nr)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")

    jetzt = datetime.now()

    writer.writerow([
        "EXTF", "700", "16", "Debitoren/Kreditoren", "5",
        jetzt.strftime("%Y%m%d%H%M%S") + "000",
        "", "", "Kanzlei AI", "",
        berater_nr, "", "", "4", "", "", "Mandanten-Stammdaten",
        "", "0", "0", "0", "EUR",
    ])

    writer.writerow([
        "Konto", "Name (Adressattyp Unternehmen)", "Unternehmensgegenstand",
        "Name (Adressattyp natürl. Person)", "Vorname (Adressattyp natürl. Person)",
        "Name (Adressattyp keine Angabe)", "Adressattyp",
        "Kurzbezeichnung", "EU-Land", "EU-UStID", "Anrede",
        "Titel/Akad. Grad", "Adelsprädikat", "Namensvorsatz", "Adressart",
        "Straße", "Postfach", "Postleitzahl", "Ort", "Land",
        "Versandzusatz", "Adresszusatz", "Abweichende Anrede",
        "Abw. Zustellbezeichnung 1", "Abw. Zustellbezeichnung 2",
        "Kennz. Korrespondenzadresse", "Adresse Gültig von", "Adresse Gültig bis",
        "Telefon", "Bemerkung (Telefon)", "Telefon GL", "Bemerkung (Telefon GL)",
        "Telefax", "Bemerkung (Telefax)", "E-Mail", "Bemerkung (E-Mail)",
        "Internet", "Bemerkung (Internet)", "IBAN-Nr. 1", "Leerfeld",
        "SWIFT-Code 1", "Abw. Kontoinhaber 1", "Kennz. Hauptbankverb. 1",
        "Bankverb 1 Gültig von", "Bankverb 1 Gültig bis",
        "Bankleitzahl 2", "Bankbezeichnung 2", "Bank-Kontonummer 2",
        "Länderkennzeichen 2", "IBAN-Nr. 2", "Leerfeld",
        "SWIFT-Code 2", "Abw. Kontoinhaber 2", "Kennz. Hauptbankverb. 2",
        "Bankverb 2 Gültig von", "Bankverb 2 Gültig bis",
        "Leerfeld", "Briefanrede", "Grußformel", "Kundennummer",
        "Steuernummer", "Sprache", "Ansprechpartner", "Vertreter",
        "Sachbearbeiter", "Diverse-Konto", "Ausgabeziel",
        "Währungssteuerung", "Kreditlimit (Debitor)",
        "Zahlungsbedingung", "Fälligkeit in Tagen (Debitor)",
        "Skonto in Prozent (Debitor)", "Kreditoren-Ziel 1 Tg.",
        "Kreditoren-Skonto 1 %", "Kreditoren-Ziel 2 Tg.",
        "Kreditoren-Skonto 2 %", "Kreditoren-Ziel 3 Brutto Tg.",
        "Kreditoren-Mindestbetrag 1", "Kreditoren-Mindestbetrag 2",
        "Zahlungsform", "Kontoinhaber", "Adressnummer des Rechnungsempfängers",
        "Umsatzsteuerpflicht", "Alternativer Suchname",
        "Status", "Anschrift manuell geändert (Felder 38-51)",
        "Anschrift individuell", "Fristberechnung bei Überweisung",
        "Mahnsperre bis", "Lastschriftsperre bis",
        "Verfalldatum Regelmäßiger Lastschriftauftrag",
        "Kunden-Typ", "Datev-Konto (OPOS)", "Leerfeld",
    ])

    idx = 0
    for name, m in mandanten.items():
        if not name or not isinstance(m, dict):
            continue
        konto = f"{10000 + idx}"
        idx += 1
        writer.writerow([
            konto,
            sanitize_datev_text(name, 50),
            m.get("branche", ""),            # Unternehmensgegenstand
            "", "",                          # Natürliche Person
            "",                              # Keine Angabe
            "2",                             # Adressattyp: Unternehmen
            sanitize_datev_text(name, 20),
            "DE",                            # EU-Land
            "",                              # EU-UStID
            "", "", "", "",                  # Anrede etc.
            "1",                             # Adressart: Straße
            "", "", "", "Deutschland", "DE", # Adresse
            "", "", "",                      # Zusätze
            "", "", "",                      # Abw. Bezeichnungen
            "", "",                          # Gültig von/bis
            m.get("telefon", ""),            # Telefon
            "", "", "",                      # Telefon weitere
            "",                              # Telefax
            "",                              # Bemerkung Telefax
            m.get("email", ""),              # E-Mail
            "", "", "",                      # Bemerkung, Internet
            "", "", "", "", "", "", "",       # IBAN 1
            "", "", "", "", "",              # Bankverb 1 Gültig
            "", "", "", "", "", "", "", "", "", "",  # Bank 2
            "", "", "",                      # Leerfeld, Briefanrede, Gruß
            "",                              # Kundennummer
            m.get("steuer_id", ""),          # Steuernummer
            "1",                             # Sprache: Deutsch
            "", "", "",                      # Ansprechpartner etc.
            "", "1", "0",                    # Div-Konto, Ausgabe, Währung
            str(round(m.get("umsatz", 0) * 0.1, 2)).replace(".", ","),  # Kreditlimit
            "14", "14", "2,00",             # Zahlungsbedingungen
            "", "", "", "", "", "", "", "",  # Kreditoren
            "", "",                          # Zahlungsform, Kontoinhaber
            "", "1",                         # Rechnung, USt-Pflicht
            sanitize_datev_text(name, 15),
            "1",                             # Status: Aktiv
            "0", "0", "0", "0",             # Manuelle Änderungen
            "", "", "",                      # Sperren
            "", "D", "", "",                # Typ, OPOS
        ])

    if idx == 0:
        raise ValueError("Keine gültigen Mandanten für DATEV-Stammdaten")

    return output.getvalue().encode(DATEV_ENCODING, errors="replace")


# ============================================================
# ELSTER-KOMPATIBLES XML
# ============================================================

def export_elster_xml(
    mandant: str,
    mandant_daten: Dict,
    steuerart: str = "UStVA",  # UStVA | ESt | KSt | GewSt
    zeitraum_jahr: int = None,
    zeitraum_quartal: int = None,
) -> bytes:
    """
    Erstellt eine ELSTER-kompatible XML-Grundstruktur.
    Format: ERiC (ELSTER Rich Client) Transfer-XML.

    Hinweis: Für echten ELSTER-Versand wird das offizielle
    ERiC-SDK von der Finanzverwaltung benötigt.
    Diese Funktion erstellt die korrekte XML-Struktur.
    """
    jahr = zeitraum_jahr or datetime.now().year
    quartal = zeitraum_quartal or ((datetime.now().month - 1) // 3 + 1)
    jetzt = datetime.now()

    # Root-Element
    root = ET.Element("Elster", {
        "xmlns": "http://www.elster.de/elsterxml/schema/v12",
        "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
    })

    # Transfer-Header
    transfer_header = ET.SubElement(root, "TransferHeader", {"version": "12"})
    ET.SubElement(transfer_header, "Verfahren").text = "ElsterAnmeldung"
    ET.SubElement(transfer_header, "DatenArt").text = steuerart
    ET.SubElement(transfer_header, "Vorgang").text = "send-Auth"
    ET.SubElement(transfer_header, "TransferTicket")
    ET.SubElement(transfer_header, "Testmerker").text = "700000004"  # Testmodus

    nutzerdaten = ET.SubElement(transfer_header, "Nutzdaten-Header", {"version": "11"})
    ET.SubElement(nutzerdaten, "NutzdatenTicket").text = "1"

    empfaenger = ET.SubElement(nutzerdaten, "Empfaenger", {"id": "F"})
    ET.SubElement(empfaenger, "Ziel").text = "9198"  # Bayerisches Finanzamt (Beispiel)

    hersteller = ET.SubElement(nutzerdaten, "HerstellerID")
    hersteller.text = "74931"  # Beispiel Hersteller-ID

    ET.SubElement(nutzerdaten, "DatenLieferant").text = "Kanzlei AI"
    ET.SubElement(nutzerdaten, "Erstellungsdatum").text = jetzt.strftime("%Y%m%d")

    # Datenteil
    datenteil = ET.SubElement(root, "DatenTeil")
    nutzdaten_block = ET.SubElement(datenteil, "Nutzdatenblock")
    nutzdaten_header = ET.SubElement(nutzdaten_block, "NutzdatenHeader", {"version": "11"})
    ET.SubElement(nutzdaten_header, "NutzdatenTicket").text = "1"

    nutzdaten = ET.SubElement(nutzdaten_block, "Nutzdaten")

    if steuerart == "UStVA":
        # Umsatzsteuer-Voranmeldung
        anmeldung = ET.SubElement(nutzdaten, "UStVA", {
            "version": "202401",
            "xmlns": "http://finkonsens.de/elster/elsterustva/ustva/v202401",
        })

        steuerpflicht = ET.SubElement(anmeldung, "Steuerpflichtiger")
        ET.SubElement(steuerpflicht, "StNr").text = mandant_daten.get("steuer_id", "0000000000")

        anmelde_inhalt = ET.SubElement(anmeldung, "UStVA")
        ET.SubElement(anmelde_inhalt, "Jahr").text = str(jahr)
        ET.SubElement(anmelde_inhalt, "Zeitraum").text = f"{quartal * 3:02d}"  # Monat

        # Umsätze (vereinfacht)
        umsatz = mandant_daten.get("umsatz", 0)
        monatsumsatz = round(umsatz / 12, 2)

        kz_81 = ET.SubElement(anmelde_inhalt, "Kz81")  # Steuerpflichtige Umsätze 19%
        kz_81.text = str(int(monatsumsatz * 100))  # In Cent

        kz_83 = ET.SubElement(anmelde_inhalt, "Kz83")  # Steuer auf Kz81
        kz_83.text = str(int(monatsumsatz * 0.19 * 100))

        kz_66 = ET.SubElement(anmelde_inhalt, "Kz66")  # Verbleibende USt
        kz_66.text = str(int(monatsumsatz * 0.19 * 100))

    elif steuerart == "GewSt":
        # Gewerbesteuer-Erklärung (vereinfacht)
        erklaerung = ET.SubElement(nutzdaten, "GewStE", {
            "version": "202301",
        })
        ET.SubElement(erklaerung, "Jahr").text = str(jahr)
        ET.SubElement(erklaerung, "StNr").text = mandant_daten.get("steuer_id", "")

    # XML serialisieren
    ET.indent(root, space="  ")
    xml_str = ET.tostring(root, encoding="unicode", xml_declaration=False)
    full_xml = '<?xml version="1.0" encoding="UTF-8"?>\n' + xml_str

    return full_xml.encode("utf-8")


# ============================================================
# EXCEL EXPORT (openpyxl)
# ============================================================

def export_excel_report(
    mandant: str,
    mandant_daten: Dict,
    aufgaben: List[Dict],
    kommunikation: List[Dict],
) -> bytes:
    """
    Erstellt einen formatierten Excel-Report für einen Mandanten.
    Enthält: Stammdaten, Aufgaben, Kommunikationshistorie, KPI-Übersicht.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl nicht installiert — pip install openpyxl")
        raise

    wb = openpyxl.Workbook()

    # ── Farben ────────────────────────────────────────────────
    DUNKEL     = "0B0D11"
    MITTEL     = "111419"
    AKZENT     = "C8A96E"
    ROT        = "E05555"
    ORANGE     = "E08C45"
    GRUEN      = "5CB87A"
    WEISS      = "E8EAF0"
    GRAU       = "8B91A0"

    def header_fill(color):
        return PatternFill("solid", fgColor=color)

    def header_font(bold=True):
        return Font(name="Calibri", bold=bold, color=WEISS, size=11)

    def cell_font(color=None, bold=False, size=10):
        return Font(name="Calibri", bold=bold,
                    color=color or "1A1A2E", size=size)

    def border_thin():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def set_col_width(ws, col, width):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ══════════════════════════════════════════════════════════
    # SHEET 1: ÜBERSICHT
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Übersicht"
    ws1.sheet_view.showGridLines = False

    # Titel
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"MANDANTEN-REPORT — {mandant.upper()}"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=18, color=AKZENT)
    ws1["A1"].fill = header_fill(DUNKEL)
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:F2")
    ws1["A2"] = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Kanzlei AI v2.0"
    ws1["A2"].font = Font(name="Calibri", size=10, color=GRAU)
    ws1["A2"].fill = header_fill(MITTEL)
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[2].height = 22

    # Stammdaten
    row = 4
    ws1[f"A{row}"] = "STAMMDATEN"
    ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color=AKZENT)
    ws1[f"A{row}"].fill = header_fill(MITTEL)
    ws1.merge_cells(f"A{row}:F{row}")
    ws1.row_dimensions[row].height = 26
    row += 1

    stammdaten = [
        ("Name",             mandant),
        ("Jahresumsatz",     f"€ {mandant_daten.get('umsatz', 0):,.2f}"),
        ("E-Mail",           mandant_daten.get("email", "—")),
        ("Telefon",          mandant_daten.get("telefon", "—")),
        ("Branche",          mandant_daten.get("branche", "—")),
        ("Steuer-ID",        mandant_daten.get("steuer_id", "—")),
        ("Notizen",          mandant_daten.get("notizen", "—")),
        ("Angelegt am",      mandant_daten.get("erstellt_am", "—")[:10]),
        ("Fehlende Docs",    str(len(mandant_daten.get("fehlende_dokumente_liste", [])))),
    ]

    for label, wert in stammdaten:
        ws1[f"A{row}"] = label
        ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=10, color=GRAU)
        ws1[f"A{row}"].fill = header_fill("F8F9FA")
        ws1[f"B{row}"] = wert
        ws1[f"B{row}"].font = Font(name="Calibri", size=10, color="1A1A2E")
        ws1.merge_cells(f"B{row}:F{row}")
        ws1.row_dimensions[row].height = 20
        row += 1

    # KPIs
    row += 1
    ws1[f"A{row}"] = "KPI ÜBERSICHT"
    ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color=AKZENT)
    ws1[f"A{row}"].fill = header_fill(MITTEL)
    ws1.merge_cells(f"A{row}:F{row}")
    ws1.row_dimensions[row].height = 26
    row += 1

    offen    = sum(1 for a in aufgaben if not a.get("erledigt"))
    erledigt = sum(1 for a in aufgaben if a.get("erledigt"))
    ueberfaellig = sum(
        1 for a in aufgaben
        if not a.get("erledigt") and a.get("tage_bis_frist", 0) is not None
        and (a.get("tage_bis_frist") or 0) < 0
    )

    kpis = [
        ("Aufgaben gesamt",  len(aufgaben),  "1A1A2E"),
        ("Aufgaben offen",   offen,          ORANGE if offen > 0 else GRUEN),
        ("Aufgaben erledigt",erledigt,       GRUEN),
        ("Überfällig",       ueberfaellig,   ROT if ueberfaellig > 0 else GRUEN),
    ]

    ws1[f"A{row}"] = "Kennzahl"
    ws1[f"B{row}"] = "Wert"
    for cell in [ws1[f"A{row}"], ws1[f"B{row}"]]:
        cell.font = Font(name="Calibri", bold=True, size=10, color=WEISS)
        cell.fill = header_fill(DUNKEL)
    row += 1

    for label, wert, farbe in kpis:
        ws1[f"A{row}"] = label
        ws1[f"B{row}"] = wert
        ws1[f"B{row}"].font = Font(name="Calibri", bold=True, size=11, color=farbe)
        ws1.row_dimensions[row].height = 20
        row += 1

    set_col_width(ws1, 1, 25)
    set_col_width(ws1, 2, 35)

    # ══════════════════════════════════════════════════════════
    # SHEET 2: AUFGABEN
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Aufgaben")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"AUFGABEN — {mandant}"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color=AKZENT)
    ws2["A1"].fill = header_fill(DUNKEL)
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 32

    headers = ["Beschreibung", "Frist", "Priorität", "Status", "Kategorie",
               "Erstellt am", "Erledigt am"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WEISS)
        cell.fill = header_fill(MITTEL)
        cell.alignment = Alignment(horizontal="center")

    widths = [40, 14, 14, 14, 20, 16, 16]
    for col, w in enumerate(widths, 1):
        set_col_width(ws2, col, w)

    for row_idx, aufgabe in enumerate(aufgaben, 3):
        erledigt = aufgabe.get("erledigt", False)
        tage     = aufgabe.get("tage_bis_frist")
        farbe    = ROT if (not erledigt and tage is not None and tage < 0) else \
                   ORANGE if (not erledigt and tage is not None and 0 <= tage <= 3) else \
                   GRUEN if erledigt else "1A1A2E"

        zeile = [
            aufgabe.get("beschreibung", ""),
            aufgabe.get("frist", ""),
            aufgabe.get("prioritaet", "normal"),
            "✓ Erledigt" if erledigt else "○ Offen",
            aufgabe.get("kategorie", ""),
            aufgabe.get("erstellt_am", "")[:10] if aufgabe.get("erstellt_am") else "",
            aufgabe.get("erledigt_am", "")[:10] if aufgabe.get("erledigt_am") else "",
        ]

        fill_farbe = "F0FFF4" if erledigt else "FFF8F8" if (tage is not None and tage < 0) else "FFFFFF"

        for col, wert in enumerate(zeile, 1):
            cell = ws2.cell(row=row_idx, column=col, value=wert)
            cell.font = Font(name="Calibri", size=10,
                             color=farbe if col == 4 else "1A1A2E")
            cell.fill = PatternFill("solid", fgColor=fill_farbe)
            cell.border = border_thin()
            if col in [2, 3, 4]:
                cell.alignment = Alignment(horizontal="center")

        ws2.row_dimensions[row_idx].height = 18

    # ══════════════════════════════════════════════════════════
    # SHEET 3: KOMMUNIKATION
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Kommunikation")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"KOMMUNIKATIONS-HISTORY — {mandant}"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color=AKZENT)
    ws3["A1"].fill = header_fill(DUNKEL)
    ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws3.row_dimensions[1].height = 32

    for col, h in enumerate(["Datum", "Typ", "Inhalt"], 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WEISS)
        cell.fill = header_fill(MITTEL)

    set_col_width(ws3, 1, 18)
    set_col_width(ws3, 2, 18)
    set_col_width(ws3, 3, 60)

    for row_idx, eintrag in enumerate(kommunikation[-50:], 3):
        try:
            dt = datetime.fromisoformat(
                eintrag.get("timestamp", eintrag.get("zeit", ""))
            ).strftime("%d.%m.%Y %H:%M")
        except Exception:
            dt = "—"

        zeile = [
            dt,
            eintrag.get("typ", "?"),
            eintrag.get("text", eintrag.get("inhalt", ""))[:200],
        ]
        for col, wert in enumerate(zeile, 1):
            cell = ws3.cell(row=row_idx, column=col, value=wert)
            cell.font = Font(name="Calibri", size=9, color="1A1A2E")
            cell.border = border_thin()
            if col == 3:
                cell.alignment = Alignment(wrap_text=True)
        ws3.row_dimensions[row_idx].height = 15

    # ══════════════════════════════════════════════════════════
    # Output als Bytes
    # ══════════════════════════════════════════════════════════
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ============================================================
# UNIVERSAL CSV EXPORT
# ============================================================

def export_csv_mandanten(mandanten: Dict[str, Dict]) -> bytes:
    """Alle Mandanten als einfache CSV für Import in andere Systeme."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)

    writer.writerow([
        "Name", "Umsatz", "Email", "Telefon", "Branche",
        "Steuer-ID", "Fehlende Dokumente", "Letzte Antwort",
        "Angelegt am", "Aktiv",
    ])

    for name, m in mandanten.items():
        writer.writerow([
            name,
            m.get("umsatz", 0),
            m.get("email", ""),
            m.get("telefon", ""),
            m.get("branche", ""),
            m.get("steuer_id", ""),
            ", ".join(m.get("fehlende_dokumente_liste", [])),
            m.get("letzte_antwort", "")[:10],
            m.get("erstellt_am", "")[:10],
            "Ja" if m.get("aktiv", True) else "Nein",
        ])

    return output.getvalue().encode("utf-8-sig")  # BOM für Excel-Kompatibilität


def export_csv_aufgaben(aufgaben: Dict[str, Dict]) -> bytes:
    """Alle Aufgaben als CSV."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)

    writer.writerow([
        "ID", "Mandant", "Beschreibung", "Frist", "Priorität",
        "Kategorie", "Status", "Erstellt am", "Erledigt am",
    ])

    for aufgabe_id, a in aufgaben.items():
        writer.writerow([
            aufgabe_id[:8],
            a.get("mandant", ""),
            a.get("beschreibung", ""),
            a.get("frist", ""),
            a.get("prioritaet", "normal"),
            a.get("kategorie", ""),
            "Erledigt" if a.get("erledigt") else "Offen",
            a.get("erstellt_am", "")[:10],
            a.get("erledigt_am", "")[:10] if a.get("erledigt_am") else "",
        ])

    return output.getvalue().encode("utf-8-sig")


# ============================================================
# ZIP-PAKET (alles gebündelt)
# ============================================================

def export_komplettpaket(
    mandant: str,
    mandant_daten: Dict,
    aufgaben_list: List[Dict],
    alle_mandanten: Dict,
    alle_aufgaben: Dict,
    kommunikation: List[Dict],
    berater_nr: str = "1234",
) -> Tuple[bytes, Dict[str, Any]]:
    """
    ZIP (ZIP_DEFLATED) mit DATEV, ELSTER, Excel, CSV, JSON + EXPORT_MANIFEST.json.
    Mindestens eine DATEV-Datei muss erfolgreich sein.
    """
    buffer = io.BytesIO()
    manifest: Dict[str, Any] = {
        "mandant": mandant,
        "exportiert_am": datetime.now().isoformat(),
        "dateien": {},
        "fehler": [],
    }
    files_written = 0

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        datum = datetime.now().strftime("%Y%m%d")
        safe_name = re.sub(r"[^\w\-]+", "_", mandant or "Mandant").strip("_") or "Mandant"
        prefix = f"{datum}_{safe_name}"

        def _add(path: str, data: bytes, key: str, extra: Optional[Dict] = None) -> None:
            nonlocal files_written
            zf.writestr(path, data)
            files_written += 1
            entry = {"pfad": path, "groesse_bytes": len(data), "status": "ok"}
            if extra:
                entry.update(extra)
            manifest["dateien"][key] = entry

        def _fail(key: str, err: Exception) -> None:
            msg = str(err)
            manifest["fehler"].append({key: msg})
            log.warning("Komplett-Export %s — %s: %s", mandant, key, msg)

        # DATEV Buchungsstapel (Pflicht für sinnvolles Paket)
        try:
            datev_buch = export_datev_buchungsstapel(
                mandant, mandant_daten, aufgaben_list, berater_nr=berater_nr
            )
            meta = validate_datev_buchungsstapel_csv(datev_buch)
            _add(
                f"DATEV/{prefix}_Buchungsstapel.csv",
                datev_buch,
                "datev_buchungsstapel",
                {"buchungen": meta.get("buchungen"), "warnings": meta.get("warnings") or []},
            )
        except Exception as e:
            _fail("datev_buchungsstapel", e)

        try:
            datev_stamm = export_datev_stammdaten(alle_mandanten, berater_nr)
            _add(f"DATEV/{datum}_Stammdaten.csv", datev_stamm, "datev_stammdaten")
        except Exception as e:
            _fail("datev_stammdaten", e)

        try:
            elster_xml = export_elster_xml(mandant, mandant_daten)
            _add(f"ELSTER/{prefix}_UStVA.xml", elster_xml, "elster_ustva")
        except Exception as e:
            _fail("elster_ustva", e)

        try:
            excel = export_excel_report(mandant, mandant_daten, aufgaben_list, kommunikation)
            _add(f"Reports/{prefix}_Report.xlsx", excel, "excel_report")
        except Exception as e:
            _fail("excel_report", e)

        try:
            mandanten_csv = export_csv_mandanten(alle_mandanten)
            _add(f"CSV/{datum}_Mandanten.csv", mandanten_csv, "csv_mandanten")
        except Exception as e:
            _fail("csv_mandanten", e)

        try:
            aufgaben_csv = export_csv_aufgaben(alle_aufgaben)
            _add(f"CSV/{datum}_Aufgaben.csv", aufgaben_csv, "csv_aufgaben")
        except Exception as e:
            _fail("csv_aufgaben", e)

        try:
            backup = {
                "mandant": mandant,
                "stammdaten": mandant_daten,
                "aufgaben": aufgaben_list,
                "exportiert_am": manifest["exportiert_am"],
            }
            _add(
                f"Backup/{prefix}_Mandant.json",
                json.dumps(backup, ensure_ascii=False, indent=2, default=str).encode("utf-8"),
                "json_backup",
            )
        except Exception as e:
            _fail("json_backup", e)

        if files_written == 0:
            raise ValueError("Komplett-Paket leer — kein Exportbestandteil konnte erstellt werden")

        if "datev_buchungsstapel" not in manifest["dateien"] and "datev_stammdaten" not in manifest["dateien"]:
            raise ValueError(
                "DATEV-Export fehlgeschlagen — ZIP ohne DATEV-Dateien. "
                "Einstellungen und Mandantendaten prüfen."
            )

        readme = f"""KANZLEI AI — EXPORT-PAKET (ZIP)
================================
Mandant:     {mandant}
Exportiert:  {datetime.now().strftime('%d.%m.%Y %H:%M')}
Dateien:     {files_written}
Fehler:      {len(manifest['fehler'])}

ORDNER:
  DATEV/     Buchungsstapel + Stammdaten (EXTF v700, Windows-1252)
  ELSTER/    UStVA XML (Grundstruktur, Test vor Versand)
  Reports/   Excel-Übersicht
  CSV/       Universal-Export
  Backup/    JSON Snapshot Mandant

DATEV IMPORT:
  1. Stammdaten importieren (Debitoren)
  2. Buchungsstapel importieren
  3. Buchungen in DATEV prüfen — Kanzlei AI ersetzt keine Fibu

Details: EXPORT_MANIFEST.json
"""
        zf.writestr("README.txt", readme.encode("utf-8"))
        manifest["dateien_gesamt"] = files_written
        zf.writestr(
            "EXPORT_MANIFEST.json",
            json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"),
        )

    buffer.seek(0)
    return buffer.read(), manifest